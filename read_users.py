import openpyxl

def inspect_excel():
    try:
        wb = openpyxl.load_workbook("User.xlsx", read_only=True)
        print("Sheets in User.xlsx:", wb.sheetnames)
        
        sheet = wb.active
        print(f"Active sheet: {sheet.title}")
        print("Rows:")
        
        # Read first 10 rows
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i >= 30:
                print("... truncated after 30 rows")
                break
            print(f"Row {i+1}: {row}")
            
    except Exception as e:
        print("Error reading Excel file:", e)

if __name__ == "__main__":
    inspect_excel()
